from __future__ import annotations

import argparse
import csv
import json
import random
import shutil
import zipfile
from datetime import date, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
RELEASES = ROOT / "releases"
random.seed(2030)

DEPARTMENTS = [
    "Operations", "Sales", "Customer Support", "Technology",
    "Finance", "Human Resources", "Supply Chain", "Marketing",
]


def ensure_dirs() -> None:
    for folder in [
        "data/csv", "analysis/notebooks", "analysis/sql", "docs", "metadata",
        "dashboards/excel", "dashboards/power-bi/dashboard-images",
        "dashboards/looker-studio", "assets", "releases",
    ]:
        (ROOT / folder).mkdir(parents=True, exist_ok=True)


def write_csv(name: str, rows: list[dict]) -> None:
    path = ROOT / "data/csv" / name
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def months() -> list[date]:
    output = []
    year, month = 2024, 8
    for index in range(24):
        y = year + (month - 1 + index) // 12
        m = (month - 1 + index) % 12 + 1
        output.append(date(y, m, 1))
    return output


def generate_csv_data() -> None:
    periods = months()
    monthly = []
    headcount = 68
    for i, period in enumerate(periods):
        progress = max(0.0, min(1.0, (i - 11) / 12))
        target = round(68 + (226 - 68) * i / 23)
        hires = max(2, round(target - headcount + random.uniform(2, 5)))
        turnover = 0.286 * (1 - progress) + 0.178 * progress
        exits = max(1, round(headcount * turnover / 12 + random.uniform(-0.5, 0.6)))
        headcount += hires - exits
        monthly.append({
            "month": period.isoformat(), "headcount": headcount, "hires": hires, "exits": exits,
            "annualized_turnover_rate": round(turnover, 4),
            "absenteeism_rate": round(0.127 * (1 - progress) + 0.076 * progress, 4),
            "overtime_cost_ratio": round(0.184 * (1 - progress) + 0.109 * progress, 4),
            "avg_engagement_score": round(56 * (1 - progress) + 74 * progress, 1),
            "avg_performance_score": round(63 * (1 - progress) + 76 * progress, 1),
            "avg_time_to_fill_days": round(53 * (1 - progress) + 32 * progress, 1),
            "training_completion_rate": round(0.44 * (1 - progress) + 0.89 * progress, 4),
            "compliance_documentation_rate": round(0.61 * (1 - progress) + 0.95 * progress, 4),
            "manager_effectiveness_score": round(58 * (1 - progress) + 77 * progress, 1),
            "early_attrition_90d_rate": round(0.241 * (1 - progress) + 0.121 * progress, 4),
            "phase": "Diagnostic / Startup" if i < 12 else "Intervention / SME Transition",
        })
    write_csv("monthly_hr_kpis.csv", monthly)

    dept_rows = []
    for i, period in enumerate(periods):
        progress = max(0.0, min(1.0, (i - 11) / 12))
        for index, dept in enumerate(DEPARTMENTS):
            bias = 1.30 - index * 0.07
            dept_rows.append({
                "month": period.isoformat(), "department": dept,
                "headcount": max(4, round(monthly[i]["headcount"] / 8 + random.uniform(-3, 3))),
                "annualized_turnover_rate": round(max(0.08, 0.286 * bias * (1 - 0.38 * progress)), 4),
                "absenteeism_rate": round(max(0.03, 0.127 * bias * (1 - 0.40 * progress)), 4),
                "avg_engagement_score": round(55 + 20 * progress + random.uniform(-3, 3), 1),
                "avg_performance_score": round(62 + 14 * progress + random.uniform(-3, 3), 1),
                "overtime_hours": round(180 * bias * (1 - 0.35 * progress), 1),
                "avg_vacancy_days": round(55 * bias * (1 - 0.35 * progress), 1),
                "people_risk_index": round(24 * bias * (1 - 0.35 * progress), 1),
            })
    write_csv("department_monthly_kpis.csv", dept_rows)

    scorecard = []
    latest = [row for row in dept_rows if row["month"] == periods[-1].isoformat()]
    for row in latest:
        risk = row["people_risk_index"]
        scorecard.append({
            "department": row["department"], "headcount": row["headcount"],
            "turnover_rate": row["annualized_turnover_rate"],
            "absenteeism_rate": row["absenteeism_rate"],
            "engagement_score": row["avg_engagement_score"],
            "performance_score": row["avg_performance_score"],
            "overtime_hours": row["overtime_hours"], "vacancy_days": row["avg_vacancy_days"],
            "people_risk_index": risk,
            "risk_band": "High" if risk >= 18 else "Medium" if risk >= 13 else "Low",
            "priority_action": "Retention and manager action plan" if risk >= 18 else "Capability and workforce planning",
        })
    write_csv("department_scorecard.csv", scorecard)

    employee_rows = []
    for i in range(1, 281):
        dept = random.choice(DEPARTMENTS)
        engagement = max(35, min(95, random.gauss(69, 11)))
        performance = max(40, min(95, random.gauss(73, 10)))
        absence = max(0.01, min(0.25, random.gauss(0.08, 0.035)))
        overtime = max(0, random.gauss(18, 8))
        risk = round((100 - engagement) * 0.35 + absence * 100 * 0.3 + max(0, 70 - performance) * 0.2 + overtime * 0.15, 1)
        employee_rows.append({
            "employee_id": f"NCS-{i:04d}", "department": dept,
            "designation_level": random.choice(["Entry", "Executive", "Senior Executive", "Assistant Manager"]),
            "location": random.choice(["Dhaka", "Gazipur", "Narayanganj", "Remote"]),
            "hire_date": (date(2023, 1, 1) + timedelta(days=random.randint(0, 1200))).isoformat(),
            "employment_status": "Active", "monthly_salary_bdt": random.randrange(24000, 85000, 500),
            "engagement_score": round(engagement, 1), "performance_score": round(performance, 1),
            "absence_rate": round(absence, 4), "monthly_overtime_hours": round(overtime, 1),
            "training_hours_last_12m": round(max(0, random.gauss(18, 7)), 1),
            "critical_role": random.choice(["Yes", "No", "No"]), "people_risk_score": risk,
            "risk_band": "High" if risk >= 32 else "Medium" if risk >= 22 else "Low",
        })
    write_csv("employees.csv", employee_rows)

    recruitment = []
    for i in range(1, 109):
        after = i > 54
        applicants = random.randint(25, 110)
        interviews = max(2, int(applicants * random.uniform(0.12, 0.25)))
        offers = max(1, int(interviews * random.uniform(0.35, 0.65)))
        joins = max(1, int(offers * random.uniform(0.72, 0.96)))
        recruitment.append({
            "requisition_id": f"REQ-{i:04d}", "department": random.choice(DEPARTMENTS),
            "source": random.choice(["Referral", "Bdjobs", "LinkedIn", "Campus", "Agency"]),
            "applicants": applicants, "interviews": interviews, "offers": offers, "joins": joins,
            "time_to_fill_days": max(14, round(random.gauss(32 if after else 53, 8))),
            "cost_per_hire_bdt": random.randrange(4000, 16000, 500),
            "quality_of_hire_score": round(random.gauss(78 if after else 64, 8), 1),
            "retention_90d_rate": round(min(1, max(0.55, random.gauss(0.88 if after else 0.76, 0.07))), 4),
            "phase": "After intervention" if after else "Before intervention",
        })
    write_csv("recruitment.csv", recruitment)

    learning = []
    programmes = ["Manager Essentials", "Structured Interviewing", "KPI and Feedback", "HR Compliance", "AI for HR"]
    for i in range(1, 401):
        before = random.randint(45, 75)
        completed = random.random() < 0.86
        learning.append({
            "learning_record_id": f"LND-{i:04d}", "employee_id": f"NCS-{random.randint(1, 280):04d}",
            "department": random.choice(DEPARTMENTS), "program": random.choice(programmes),
            "hours": random.choice([2, 4, 6, 8, 12]), "completion_status": "Yes" if completed else "No",
            "pre_assessment_score": before, "post_assessment_score": min(100, before + random.randint(8, 24)),
            "behavior_application_score": random.randint(55, 92),
        })
    write_csv("learning.csv", learning)

    interventions = []
    areas = ["Talent acquisition", "Learning and development", "Performance management", "Culture",
             "Compliance and risk", "Data-driven decisions", "Change management", "DEI", "Wellbeing", "AI integration"]
    for i, area in enumerate(areas, 1):
        interventions.append({
            "strategy_area": f"{i}. {area}", "business_problem": "Growth-related people risk",
            "intervention": f"Structured {area.lower()} operating practice",
            "baseline": 50 + i, "target": 75 + i, "actual": 72 + i,
            "owner": "HR and Business", "status": "Completed" if i < 9 else "Pilot",
        })
    write_csv("intervention_impact.csv", interventions)

    dictionary = []
    for filename in ["employees.csv", "monthly_hr_kpis.csv", "department_scorecard.csv", "recruitment.csv", "learning.csv"]:
        with (ROOT / "data/csv" / filename).open(encoding="utf-8-sig") as fh:
            fields = next(csv.reader(fh))
        for field in fields:
            dictionary.append({"file": filename, "field": field, "data_type": "synthetic", "description": field.replace("_", " ").title()})
    write_csv("data_dictionary.csv", dictionary)


def write_core_docs(version: str) -> None:
    (ROOT / "VERSION").write_text(version + "\n", encoding="utf-8")
    (ROOT / "requirements.txt").write_text("pandas>=2.0\nmatplotlib>=3.7\nopenpyxl>=3.1\n", encoding="utf-8")
    (ROOT / "docs/CASE_STUDY.md").write_text("""# Case Study\n\nA synthetic Bangladesh startup is becoming an SME while turnover, absence, overtime, recruitment delay, manager capability and documentation risk threaten growth. The HR Executive diagnoses the connected problem, introduces accountable interventions and tracks business-facing outcomes.\n""", encoding="utf-8")
    (ROOT / "docs/PROMOTION_PORTFOLIO.md").write_text("""# Promotion Portfolio\n\nThis project demonstrates readiness to progress from HR Executive to Senior Executive or entry mid-level strategic HR work through business diagnosis, analytics, stakeholder ownership, change management, governance and executive communication.\n""", encoding="utf-8")
    (ROOT / "docs/PROJECT_USAGE_GUIDE.md").write_text("""# Project Usage Guide\n\n1. Load files from `data/csv/`.\n2. Open the Excel dashboard.\n3. Rebuild the supplied Power BI SVG concepts using the model and DAX guide.\n4. Use the Looker Studio specification for a shareable report.\n5. Run the notebooks and SQL for practice.\n""", encoding="utf-8")
    (ROOT / "docs/ETHICS_AND_LIMITATIONS.md").write_text("""# Ethics and Limitations\n\nAll data is synthetic. Risk scores support discussion only and must never automate employment decisions. Before/after movement is descriptive and does not prove causality.\n""", encoding="utf-8")
    (ROOT / "analysis/sql/hr_strategy_queries.sql").write_text("""SELECT department, people_risk_index, risk_band FROM department_scorecard ORDER BY people_risk_index DESC;\n\nSELECT phase, AVG(time_to_fill_days) AS avg_time_to_fill FROM recruitment GROUP BY phase;\n""", encoding="utf-8")
    notebook = {"cells": [{"cell_type": "markdown", "metadata": {}, "source": ["# HR Diagnostic\n", "Synthetic portfolio analysis."]}, {"cell_type": "code", "execution_count": None, "metadata": {}, "outputs": [], "source": ["import pandas as pd\n", "monthly = pd.read_csv('../../data/csv/monthly_hr_kpis.csv')\n", "monthly.head()\n"]}], "metadata": {"kernelspec": {"display_name": "Python 3", "language": "python", "name": "python3"}}, "nbformat": 4, "nbformat_minor": 5}
    (ROOT / "analysis/notebooks/01_hr_diagnostic.ipynb").write_text(json.dumps(notebook, indent=2), encoding="utf-8")
    notebook["cells"][0]["source"] = ["# Intervention Impact\n", "Descriptive before/after analysis."]
    notebook["cells"][1]["source"] = ["import pandas as pd\n", "monthly = pd.read_csv('../../data/csv/monthly_hr_kpis.csv')\n", "monthly.groupby('phase').mean(numeric_only=True)\n"]
    (ROOT / "analysis/notebooks/02_intervention_impact.ipynb").write_text(json.dumps(notebook, indent=2), encoding="utf-8")
    metadata = {
        "project_name": "Bangladesh SME HR Strategy Transformation", "repository": "samusa099/bd-hr-analytics",
        "version": version, "portfolio_owner": "Musa", "case_type": "Synthetic portfolio simulation",
        "tools": ["Excel", "Power BI", "Looker Studio", "Python", "SQL"],
        "data_classification": "Synthetic; no personal or confidential data",
    }
    (ROOT / "metadata/project_metadata.json").write_text(json.dumps(metadata, indent=2), encoding="utf-8")
    kaggle = {"title": "Bangladesh SME HR Strategy Transformation", "id": "samusa099/bd-hr-analytics", "licenses": [{"name": "CC-BY-4.0"}], "subtitle": "Synthetic startup-to-SME people analytics portfolio", "description": "Synthetic Bangladesh HR analytics portfolio dataset.", "keywords": ["human resources", "people analytics", "Bangladesh", "Power BI", "Excel"]}
    (ROOT / "metadata/dataset-metadata.json").write_text(json.dumps(kaggle, indent=2), encoding="utf-8")


def svg_dashboard(title: str, number: int) -> str:
    colors = ["#2563EB", "#0D9488", "#F97316", "#7C3AED", "#0891B2", "#16A34A"]
    cards = []
    for i in range(6):
        x = 270 + i * 260
        cards.append(f'<rect x="{x}" y="205" width="240" height="130" rx="18" fill="#fff" stroke="#E2E8F0"/><circle cx="{x+38}" cy="250" r="24" fill="{colors[i]}"/><text x="{x+72}" y="242" font-family="Arial" font-size="14" fill="#475569">KPI {i+1}</text><text x="{x+72}" y="285" font-family="Arial" font-size="28" font-weight="700" fill="#0F172A">{72+i*3}</text>')
    panels = []
    for i, (x, y, w, h) in enumerate([(270,370,750,285),(1045,370,800,285),(270,680,750,300),(1045,680,800,300)]):
        panels.append(f'<rect x="{x}" y="{y}" width="{w}" height="{h}" rx="18" fill="#fff" stroke="#E2E8F0"/><text x="{x+22}" y="{y+38}" font-family="Arial" font-size="20" font-weight="700" fill="#1E3A8A">Analysis Panel {i+1}</text>')
        for j in range(6):
            bw = 420 - j * 45
            panels.append(f'<rect x="{x+150}" y="{y+65+j*30}" width="{bw}" height="18" rx="5" fill="{colors[j]}" opacity="0.85"/>')
    return f'''<svg xmlns="http://www.w3.org/2000/svg" width="1920" height="1080" viewBox="0 0 1920 1080"><rect width="1920" height="1080" fill="#F8FAFC"/><rect width="1920" height="90" fill="#fff"/><rect y="90" width="235" height="990" fill="#fff" stroke="#E2E8F0"/><text x="125" y="45" font-family="Arial" font-size="28" font-weight="700" fill="#0F172A">Bangladesh SME HR Strategy Transformation</text><text x="270" y="145" font-family="Arial" font-size="34" font-weight="700" fill="#0F172A">{title}</text><text x="270" y="178" font-family="Arial" font-size="17" fill="#64748B">Synthetic demo data for portfolio practice</text>{''.join(cards)}{''.join(panels)}<text x="270" y="1040" font-family="Arial" font-size="16" fill="#2563EB">Vector SVG dashboard concept — sharp at any zoom level</text></svg>'''


def write_v03_assets() -> None:
    titles = ["Executive Overview", "Workforce & Turnover", "Talent Acquisition", "Learning & Manager Capability", "Performance & Employee Experience", "Compliance, Risk & Strategic Initiatives"]
    for i, title in enumerate(titles, 1):
        slug = title.lower().replace(" & ", "-").replace(", ", "-").replace(" ", "-")
        (ROOT / f"dashboards/power-bi/dashboard-images/{i:02d}-{slug}.svg").write_text(svg_dashboard(title, i), encoding="utf-8")
    (ROOT / "dashboards/power-bi/README.md").write_text("# Power BI Build Guide\n\nUse the CSV folder, star schema, supplied DAX and six scalable SVG report-page concepts.\n", encoding="utf-8")
    (ROOT / "dashboards/power-bi/data_model.md").write_text("# Data Model\n\nDimDate and DimDepartment filter monthly KPI, department, recruitment and learning fact tables using one-to-many relationships.\n", encoding="utf-8")
    (ROOT / "dashboards/power-bi/measures.dax").write_text("Headcount = MAX(FactMonthlyKPI[headcount])\nTurnover Rate = AVERAGE(FactMonthlyKPI[annualized_turnover_rate])\nAbsenteeism Rate = AVERAGE(FactMonthlyKPI[absenteeism_rate])\nAverage Engagement = AVERAGE(FactMonthlyKPI[avg_engagement_score])\n", encoding="utf-8")
    (ROOT / "dashboards/power-bi/theme.json").write_text(json.dumps({"name": "BD HR Analytics", "dataColors": ["#2563EB", "#14B8A6", "#F97316", "#7C3AED"]}, indent=2), encoding="utf-8")
    (ROOT / "dashboards/looker-studio/README.md").write_text("# Looker Studio Guide\n\nConnect the CSV files and build leadership, department risk, recruitment, learning and intervention pages.\n", encoding="utf-8")
    (ROOT / "dashboards/looker-studio/calculated_fields.md").write_text("# Calculated Fields\n\n`SUM(hires) - SUM(exits)`\n\n`AVG(annualized_turnover_rate)`\n", encoding="utf-8")
    (ROOT / "assets/project-cover.svg").write_text(svg_dashboard("HR Strategy Analytics Portfolio", 0), encoding="utf-8")
    create_excel()


def create_excel() -> None:
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, BarChart, Reference
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Executive Dashboard"
    ws.merge_cells("A1:L2")
    ws["A1"] = "Bangladesh SME HR Strategy Transformation"
    ws["A1"].font = Font(size=20, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0F172A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    kpis = [("Turnover", "17.8%"), ("Absenteeism", "7.6%"), ("Time to Fill", "32 days"), ("Engagement", "74/100"), ("Compliance", "95%"), ("Manager Effectiveness", "77/100")]
    for i, (label, value) in enumerate(kpis):
        col = 1 + i * 2
        ws.cell(4, col, label)
        ws.cell(5, col, value)
        ws.cell(4, col).font = Font(bold=True, color="FFFFFF")
        ws.cell(4, col).fill = PatternFill("solid", fgColor="2563EB")
        ws.cell(5, col).font = Font(size=16, bold=True)
    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 16

    trend = wb.create_sheet("KPI Trend")
    with (ROOT / "data/csv/monthly_hr_kpis.csv").open(encoding="utf-8-sig") as fh:
        rows = list(csv.DictReader(fh))
    headers = list(rows[0])
    trend.append(headers)
    for row in rows:
        trend.append([row[h] for h in headers])
    chart = LineChart()
    chart.title = "Turnover and Absenteeism Trend"
    chart.add_data(Reference(trend, min_col=5, max_col=6, min_row=1, max_row=len(rows)+1), titles_from_data=True)
    chart.set_categories(Reference(trend, min_col=1, min_row=2, max_row=len(rows)+1))
    ws.add_chart(chart, "A8")

    dept = wb.create_sheet("Department Scorecard")
    with (ROOT / "data/csv/department_scorecard.csv").open(encoding="utf-8-sig") as fh:
        drows = list(csv.DictReader(fh))
    dept.append(list(drows[0]))
    for row in drows:
        dept.append(list(row.values()))
    bar = BarChart()
    bar.title = "People Risk by Department"
    bar.add_data(Reference(dept, min_col=9, min_row=1, max_row=len(drows)+1), titles_from_data=True)
    bar.set_categories(Reference(dept, min_col=1, min_row=2, max_row=len(drows)+1))
    ws.add_chart(bar, "G8")

    wb.save(ROOT / "dashboards/excel/HR_Strategy_Transformation_Dashboard.xlsx")


def update_readme() -> None:
    path = ROOT / "README.md"
    text = path.read_text(encoding="utf-8") if path.exists() else "# Bangladesh SME HR Strategy Transformation\n"
    marker = "## 📥 Release Downloads"
    if marker not in text:
        text += "\n\n## 📥 Release Downloads\n\n- [Project Version v0.1.0](releases/bd-hr-analytics-v0.1.0.zip)\n- [Unified ZIP v0.3.0](releases/bd-hr-analytics-unified-v0.3.0.zip)\n- [Latest unified project ZIP](bd-hr-analytics-unified-project.zip)\n"
    path.write_text(text, encoding="utf-8")


def write_manifest() -> None:
    files = []
    for path in sorted(ROOT.rglob("*")):
        if path.is_file() and ".git" not in path.parts and "releases" not in path.parts and path.name != "bd-hr-analytics-unified-project.zip":
            files.append({"path": str(path.relative_to(ROOT)), "size_bytes": path.stat().st_size})
    (ROOT / "metadata/file_manifest.json").write_text(json.dumps(files, indent=2), encoding="utf-8")


def make_zip(destination: Path, include_dashboards: bool) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    if destination.exists():
        destination.unlink()
    allowed = ["README.md", "VERSION", "requirements.txt", "data", "analysis", "docs", "metadata"]
    if include_dashboards:
        allowed += ["dashboards", "assets"]
    with zipfile.ZipFile(destination, "w", zipfile.ZIP_DEFLATED, compresslevel=9) as archive:
        for item in allowed:
            path = ROOT / item
            if path.is_file():
                archive.write(path, Path("bd-hr-analytics") / item)
            elif path.is_dir():
                for file in sorted(path.rglob("*")):
                    if file.is_file():
                        archive.write(file, Path("bd-hr-analytics") / file.relative_to(ROOT))


def build(phase: str) -> None:
    ensure_dirs()
    generate_csv_data()
    if phase == "v0.1":
        write_core_docs("v0.1.0")
        update_readme()
        write_manifest()
        make_zip(RELEASES / "bd-hr-analytics-v0.1.0.zip", include_dashboards=False)
    elif phase == "v0.3":
        write_core_docs("v0.3.0")
        write_v03_assets()
        update_readme()
        write_manifest()
        make_zip(RELEASES / "bd-hr-analytics-unified-v0.3.0.zip", include_dashboards=True)
    else:
        raise ValueError(f"Unknown phase: {phase}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--phase", choices=["v0.1", "v0.3"], required=True)
    build(parser.parse_args().phase)
